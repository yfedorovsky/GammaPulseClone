"""TraderMir commons portfolio → Excel spreadsheet.

Reads the parsed events CSV (from trader_mir_commons_parser_v3.py),
builds the per-ticker position ledger, enriches with live prices via
Tradier, computes SPY benchmark for alpha comparison, and emits an
xlsx with two tabs:

  Tab 1: Current Positions  — open positions with cost basis, current
                              price, unrealized P&L, vs-SPY alpha since
                              first entry
  Tab 2: Trade History      — all events + per-ticker realized P&L on
                              closed positions + closed-vs-SPY alpha

Sizing model: TraderMir doesn't post share counts, so we use a UNIT-
based tracker where each OPEN/ADD adds 1 unit. TRIM = -0.5 units.
CLOSE = -all remaining units. Cost basis is the AVG of entry prices
weighted by units. This produces a clean trade-level P&L % comparable
across positions even without knowing actual share count.

Output: discord/trader_mir_commons_portfolio_<DATE>.xlsx
"""
from __future__ import annotations

import csv
import os
import sys
from collections import defaultdict, deque
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from server.config import get_settings

EVENTS_CSV = ROOT / "discord" / "commons_parsed_events_v3.csv"
TODAY = date.today().isoformat()
OUT_XLSX = ROOT / "discord" / f"trader_mir_commons_portfolio_{TODAY}_v4.xlsx"

TRADIER_TOKEN = (
    os.environ.get("TRADIER_TOKEN")
    or os.environ.get("TRADIER_API_TOKEN")
    or get_settings().tradier_token
)
TRADIER_BASE = "https://api.tradier.com/v1"


# ── Data fetchers ──────────────────────────────────────────────────────

def fetch_current_quote(client: httpx.Client, ticker: str) -> float | None:
    """Tradier live quote."""
    try:
        r = client.get(
            f"{TRADIER_BASE}/markets/quotes",
            params={"symbols": ticker},
            headers={"Authorization": f"Bearer {TRADIER_TOKEN}", "Accept": "application/json"},
            timeout=8.0,
        )
        if r.status_code != 200:
            return None
        q = r.json().get("quotes", {}).get("quote", {})
        if isinstance(q, list):
            q = q[0] if q else {}
        return float(q.get("last") or q.get("close") or 0) or None
    except Exception as e:
        print(f"  [warn] {ticker} quote failed: {e}", file=sys.stderr)
        return None


def fetch_close_on_date(client: httpx.Client, ticker: str, d: str) -> float | None:
    """SPY/ticker close on a specific date. d = YYYY-MM-DD."""
    try:
        r = client.get(
            f"{TRADIER_BASE}/markets/history",
            params={"symbol": ticker, "interval": "daily", "start": d,
                    "end": (datetime.strptime(d, "%Y-%m-%d").date() + timedelta(days=10)).isoformat()},
            headers={"Authorization": f"Bearer {TRADIER_TOKEN}", "Accept": "application/json"},
            timeout=8.0,
        )
        if r.status_code != 200:
            return None
        bars = r.json().get("history", {}).get("day") or []
        if isinstance(bars, dict):
            bars = [bars]
        # Find first bar on or after `d`
        for b in bars:
            if b.get("date") and b.get("close"):
                if b["date"] >= d:
                    return float(b["close"])
        return None
    except Exception as e:
        print(f"  [warn] {ticker} history on {d} failed: {e}", file=sys.stderr)
        return None


# ── Position ledger ────────────────────────────────────────────────────

def build_ledger(events: list[dict]) -> dict[str, dict[str, Any]]:
    """Walk events chronologically per ticker, build position state.

    Returns: ticker -> {
        events: [...],
        units_open: float,
        avg_cost: float,
        realized_pnl_per_unit: float (sum of all closed trades' P/L per original unit),
        first_entry_date: str,
        last_exit_date: str | None  (None if still open),
        status: 'OPEN' | 'CLOSED',
        entries: list of (date, price) for adds,
        exits: list of (date, price, fraction_closed) for trims/closes,
    }
    """
    by_ticker: dict[str, list[dict]] = defaultdict(list)
    for e in events:
        by_ticker[e["ticker"]].append(e)

    ledger: dict[str, dict[str, Any]] = {}
    for ticker, evs in by_ticker.items():
        evs.sort(key=lambda x: x["timestamp"])
        units = 0.0
        entries: list[tuple[str, float]] = []   # (date, price)
        exits: list[tuple[str, float, float]] = []  # (date, price, fraction)
        realized_per_unit = 0.0
        first_entry_date = None
        last_exit_date = None
        peak_units = 0.0   # max units ever held (for fraction sizing)

        for e in evs:
            d = e["date"]
            p = e["price"]
            action = e["action"]

            if action in ("OPEN", "ADD"):
                if p is None:
                    continue
                # OPEN restarts position if currently flat; ADD is +1 unit
                if action == "OPEN" and units == 0:
                    units = 1.0
                    entries.append((d, p))
                    peak_units = max(peak_units, units)
                    if first_entry_date is None:
                        first_entry_date = d
                else:
                    units += 1.0
                    entries.append((d, p))
                    peak_units = max(peak_units, units)
                    if first_entry_date is None:
                        first_entry_date = d
            elif action == "TRIM":
                if units <= 0 or p is None:
                    continue
                # Trim 50% of CURRENT units
                sold = units * 0.5
                if entries:
                    avg_cost = sum(price for _, price in entries) / len(entries)
                    realized_per_unit += (p - avg_cost) * (sold / peak_units)
                units -= sold
                exits.append((d, p, sold / peak_units))
                last_exit_date = d
            elif action in ("CLOSE", "STOP"):
                if units <= 0 or p is None:
                    continue
                sold = units
                if entries:
                    avg_cost = sum(price for _, price in entries) / len(entries)
                    realized_per_unit += (p - avg_cost) * (sold / peak_units)
                units = 0.0
                # Reset entries because position closed; if re-opened later
                # the new position has its own cost basis. Keep them in
                # history list but track current basis from a fresh entries
                # list going forward.
                entries = []
                exits.append((d, p, sold / peak_units if peak_units > 0 else 1.0))
                last_exit_date = d
                peak_units = 0.0

        # Average cost is mean of remaining entries (if still open)
        avg_cost_open = sum(price for _, price in entries) / len(entries) if entries else 0.0

        ledger[ticker] = {
            "events": evs,
            "units_open": units,
            "avg_cost": avg_cost_open,
            "realized_per_unit": realized_per_unit,
            "first_entry_date": first_entry_date,
            "last_exit_date": last_exit_date if units == 0 else None,
            "status": "OPEN" if units > 0 else "CLOSED",
            "entries": entries,
            "exits": [(d, p, frac) for d, p, frac in
                      [(e[0], e[1], e[2]) for e in
                       [(ev["date"], ev["price"], 0) for ev in evs if ev["action"] in ("TRIM", "CLOSE", "STOP")]]
                      if d],  # historical exits for reference
        }
    return ledger


# ── Spreadsheet generation ─────────────────────────────────────────────

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
BODY_FONT = Font(name="Arial")
GAIN_FILL = PatternFill("solid", start_color="C6EFCE")
LOSS_FILL = PatternFill("solid", start_color="FFC7CE")


def fmt_pct(v: float | None) -> str:
    return f"{v:+.1f}%" if v is not None else "—"


def fmt_money(v: float | None) -> str:
    return f"${v:.2f}" if v is not None else "—"


def style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def build_notes(events: list[dict]) -> str:
    """Build a compact human-readable trade history for the Notes column.

    Format: "Entries: 9/17 OPEN $44.75, 10/17 ADD $50. Exits: 3/2 TRIM $44.75,
    5/1 TRIM $98.60 (+120%)."
    """
    if not events:
        return ""
    entries = [e for e in events if e["action"] in ("OPEN", "ADD") and e.get("price")]
    exits = [e for e in events if e["action"] in ("TRIM", "CLOSE", "STOP") and e.get("price")]

    def short_date(d: str) -> str:
        # YYYY-MM-DD → M/D
        try:
            parts = d.split("-")
            return f"{int(parts[1])}/{int(parts[2])}"
        except Exception:
            return d

    parts = []
    if entries:
        n = len(entries)
        bits = [f"{short_date(e['date'])} {e['action']} ${e['price']:.2f}" for e in entries]
        # Compute avg entry % from first entry to add P&L context per add
        first_price = entries[0]["price"]
        if n > 1:
            bits_with_delta = []
            for e in entries:
                d = (e["price"] - first_price) / first_price * 100 if first_price else 0
                if abs(d) < 0.5:
                    bits_with_delta.append(f"{short_date(e['date'])} {e['action']} ${e['price']:.2f}")
                else:
                    bits_with_delta.append(
                        f"{short_date(e['date'])} {e['action']} ${e['price']:.2f} ({d:+.0f}% from initial)")
            bits = bits_with_delta
        parts.append(f"Entries ({n}x): " + "; ".join(bits))
    if exits:
        bits = []
        for e in exits:
            d = short_date(e["date"])
            act = e["action"]
            price = e["price"]
            # Compute gain vs first entry
            if entries:
                first = entries[0]["price"]
                gain = (price - first) / first * 100 if first else 0
                bits.append(f"{d} {act} ${price:.2f} ({gain:+.0f}%)")
            else:
                bits.append(f"{d} {act} ${price:.2f}")
        parts.append(f"Exits ({len(exits)}x): " + "; ".join(bits))
    return ". ".join(parts)


def build_spreadsheet(ledger: dict[str, dict[str, Any]],
                       quotes: dict[str, float],
                       spy_entry: dict[str, float],
                       spy_exit: dict[str, float],
                       spy_current: float) -> None:
    wb = Workbook()

    # ── Tab 1: Current Positions ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Current Positions"
    headers = ["Ticker", "Status", "First Entry", "Days Held", "Avg Cost",
               "Current Price", "Unrealized %", "SPY Entry", "SPY Current",
               "SPY % Move", "Alpha vs SPY", "# Entries", "# Exits", "Notes"]
    ws1.append(headers)
    style_header(ws1, 1, len(headers))

    open_positions = sorted(
        [(t, st) for t, st in ledger.items() if st["status"] == "OPEN"],
        key=lambda kv: kv[0],
    )

    for ticker, st in open_positions:
        cur = quotes.get(ticker)
        avg_cost = st["avg_cost"]
        unreal_pct = (cur - avg_cost) / avg_cost * 100 if cur and avg_cost else None
        first_entry = st["first_entry_date"]
        days_held = (date.today() - datetime.strptime(first_entry, "%Y-%m-%d").date()).days if first_entry else None
        spy_in = spy_entry.get(first_entry) if first_entry else None
        spy_move = (spy_current - spy_in) / spy_in * 100 if spy_in else None
        alpha = (unreal_pct - spy_move) if (unreal_pct is not None and spy_move is not None) else None

        notes = build_notes(st["events"])
        row = [
            ticker,
            "OPEN",
            first_entry,
            days_held,
            avg_cost if avg_cost else None,
            cur,
            unreal_pct / 100 if unreal_pct is not None else None,
            spy_in,
            spy_current,
            spy_move / 100 if spy_move is not None else None,
            alpha / 100 if alpha is not None else None,
            len(st["entries"]),
            len([e for e in st["events"] if e["action"] in ("TRIM", "CLOSE", "STOP")]),
            notes,
        ]
        ws1.append(row)

        # Color cells by P&L sign
        r = ws1.max_row
        for col_idx, val in [(7, unreal_pct), (11, alpha)]:
            cell = ws1.cell(row=r, column=col_idx)
            if val is not None:
                cell.fill = GAIN_FILL if val >= 0 else LOSS_FILL

    # Format numeric columns
    for r in range(2, ws1.max_row + 1):
        ws1.cell(row=r, column=5).number_format = '$#,##0.00'  # Avg Cost
        ws1.cell(row=r, column=6).number_format = '$#,##0.00'  # Current
        ws1.cell(row=r, column=7).number_format = '0.0%'        # Unreal %
        ws1.cell(row=r, column=8).number_format = '$#,##0.00'  # SPY Entry
        ws1.cell(row=r, column=9).number_format = '$#,##0.00'  # SPY Current
        ws1.cell(row=r, column=10).number_format = '0.0%'       # SPY Move
        ws1.cell(row=r, column=11).number_format = '0.0%'       # Alpha
        for c in range(1, len(headers) + 1):
            ws1.cell(row=r, column=c).font = BODY_FONT
            ws1.cell(row=r, column=c).border = BORDER

    # Column widths
    widths = [8, 7, 12, 10, 10, 12, 12, 10, 12, 12, 13, 9, 8, 100]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    # Wrap text + top-align Notes column
    for r in range(2, ws1.max_row + 1):
        ws1.cell(row=r, column=14).alignment = Alignment(vertical="top", wrap_text=True)
    ws1.freeze_panes = "A2"

    # ── Tab 2: Trade History + Closed P&L ───────────────────────────
    ws2 = wb.create_sheet("Trade History")

    # Top section: per-ticker closed summary
    ws2.append(["TICKER SUMMARY — Realized P&L (closed positions only)"])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14, name="Arial")
    ws2.append([])
    sum_headers = ["Ticker", "Status", "First Entry", "Last Exit", "Days Held",
                   "Avg Entry", "Avg Exit", "Realized %", "SPY Entry", "SPY Exit",
                   "SPY %", "Alpha", "# Adds", "# Exits", "Notes"]
    ws2.append(sum_headers)
    style_header(ws2, 3, len(sum_headers))

    closed_positions = sorted(
        [(t, st) for t, st in ledger.items() if st["status"] == "CLOSED"],
        key=lambda kv: kv[1]["last_exit_date"] or "",
        reverse=True,
    )

    for ticker, st in closed_positions:
        # Compute simple realized % using avg of all entry prices vs avg of all exit prices
        entry_evs = [e for e in st["events"] if e["action"] in ("OPEN", "ADD") and e["price"]]
        exit_evs = [e for e in st["events"] if e["action"] in ("TRIM", "CLOSE", "STOP") and e["price"]]
        avg_entry = sum(e["price"] for e in entry_evs) / len(entry_evs) if entry_evs else None
        avg_exit = sum(e["price"] for e in exit_evs) / len(exit_evs) if exit_evs else None
        realized = (avg_exit - avg_entry) / avg_entry * 100 if (avg_entry and avg_exit) else None

        first_entry = st["first_entry_date"]
        last_exit = st["last_exit_date"]
        days_held = ((datetime.strptime(last_exit, "%Y-%m-%d").date() -
                       datetime.strptime(first_entry, "%Y-%m-%d").date()).days
                      if first_entry and last_exit else None)

        spy_in = spy_entry.get(first_entry) if first_entry else None
        spy_out = spy_exit.get(last_exit) if last_exit else None
        spy_move = (spy_out - spy_in) / spy_in * 100 if (spy_in and spy_out) else None
        alpha = (realized - spy_move) if (realized is not None and spy_move is not None) else None

        notes_closed = build_notes(st["events"])
        row = [
            ticker, "CLOSED", first_entry, last_exit, days_held,
            avg_entry, avg_exit,
            realized / 100 if realized is not None else None,
            spy_in, spy_out,
            spy_move / 100 if spy_move is not None else None,
            alpha / 100 if alpha is not None else None,
            len(entry_evs), len(exit_evs),
            notes_closed,
        ]
        ws2.append(row)

        r = ws2.max_row
        for col_idx, val in [(8, realized), (12, alpha)]:
            cell = ws2.cell(row=r, column=col_idx)
            if val is not None:
                cell.fill = GAIN_FILL if val >= 0 else LOSS_FILL

    # Format closed summary numeric cols
    for r in range(4, ws2.max_row + 1):
        for cidx in [6, 7, 9, 10]:
            ws2.cell(row=r, column=cidx).number_format = '$#,##0.00'
        for cidx in [8, 11, 12]:
            ws2.cell(row=r, column=cidx).number_format = '0.0%'
        for c in range(1, len(sum_headers) + 1):
            ws2.cell(row=r, column=c).font = BODY_FONT
            ws2.cell(row=r, column=c).border = BORDER

    # Add a separator + bottom section: raw event log
    ws2.append([])
    ws2.append(["RAW EVENT LOG (chronological)"])
    sep_row = ws2.max_row
    ws2.cell(row=sep_row, column=1).font = Font(bold=True, size=14, name="Arial")
    ws2.append([])
    raw_headers = ["Date", "Ticker", "Action", "Price", "Tag", "Source Snippet"]
    ws2.append(raw_headers)
    style_header(ws2, ws2.max_row, len(raw_headers))

    # Flatten events sorted by ticker then time
    all_events = []
    for ticker, st in ledger.items():
        all_events.extend(st["events"])
    all_events.sort(key=lambda e: e["timestamp"])
    for e in all_events:
        ws2.append([e["date"], e["ticker"], e["action"], e.get("price"),
                    e.get("tag", ""), e.get("raw_snippet", "")[:150]])

    # Format
    for r in range(sep_row + 3, ws2.max_row + 1):
        if r > sep_row + 3:
            ws2.cell(row=r, column=4).number_format = '$#,##0.00'
        for c in range(1, len(raw_headers) + 1):
            ws2.cell(row=r, column=c).font = BODY_FONT
            ws2.cell(row=r, column=c).border = BORDER
            ws2.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=False)

    # Column widths for ws2
    widths2 = [12, 8, 12, 12, 10, 12, 12, 11, 11, 11, 9, 9, 8, 8, 100]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    # Wrap Notes column (col 15) for closed-summary rows
    for r in range(4, len(closed_positions) + 4):
        ws2.cell(row=r, column=15).alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUT_XLSX)
    print(f"Wrote spreadsheet to {OUT_XLSX}", file=sys.stderr)


# ── Main ───────────────────────────────────────────────────────────────

def apply_manual_overrides(events: list[dict]) -> list[dict]:
    """Apply operator-curated corrections from commons_manual_overrides.csv.

    Each override row has an `operation` field:
      UPDATE_ACTION_FROM_X     change matched event's action (and price if specified)
      UPDATE_PRICE_FROM_X      change matched event's price
      INSERT_NEW               add a new event not in the auto-parsed set
      INSERT_NEW_WITH_FALLBACK same as INSERT_NEW but with caveat note
    """
    override_path = ROOT / "discord" / "commons_manual_overrides.csv"
    if not override_path.exists():
        return events
    with override_path.open("r", encoding="utf-8") as f:
        overrides = list(csv.DictReader(f))

    applied = 0
    for ov in overrides:
        date = ov["date"]
        ticker = ov["ticker"]
        new_action = ov["action"]
        new_price = float(ov["price"]) if ov["price"] else None
        op = ov["operation"]

        if op.startswith("UPDATE"):
            # Determine the EXPECTED current action from the operation suffix
            # (e.g. UPDATE_ACTION_FROM_CLOSE → find an event currently tagged CLOSE)
            expected_current: str | None = None
            if "FROM_" in op:
                expected_current = op.split("FROM_")[-1].strip()
            for e in events:
                if e["date"] != date or e["ticker"] != ticker:
                    continue
                if expected_current and e["action"] != expected_current:
                    continue
                e["action"] = new_action
                if new_price is not None:
                    e["price"] = new_price
                e["tag"] = (e.get("tag", "") + " OVERRIDDEN").strip()
                applied += 1
                break
        elif op.startswith("DELETE"):
            # Remove a spurious event. Match on (date, ticker, current_action, price).
            # Used for cross-ticker phantom events like the MP CLOSE @ $6.36 that
            # got generated from the NB recap line "$NB - $8.83 (closing @ $6.36
            # - replacing with $MP)".
            expected_current = None
            if "_FROM_" in op:
                expected_current = op.split("_FROM_")[-1].strip()
            for i, e in enumerate(events):
                if e["date"] != date or e["ticker"] != ticker:
                    continue
                if expected_current and e["action"] != expected_current:
                    continue
                if new_price is not None and abs((e.get("price") or 0) - new_price) > 0.01:
                    continue
                events.pop(i)
                applied += 1
                break
        elif op.startswith("INSERT"):
            # Use T23:59:59 so override sorts AFTER all auto-parsed events
            # on the same date. Important for sequence-sensitive cases like
            # AMPX 5/21 (stop hit, then re-add) where the override must
            # come AFTER the auto-parsed STOP event in the ledger walk.
            events.append({
                "date": date,
                "timestamp": date + "T23:59:59",
                "ticker": ticker,
                "action": new_action,
                "price": new_price,
                "tag": "OVERRIDE_INSERT",
                "raw_snippet": ov["note"],
                "msg_id": "manual_override",
            })
            applied += 1

    events.sort(key=lambda e: e["timestamp"])
    print(f"Applied {applied} manual overrides from {override_path.name}",
          file=sys.stderr)
    return events


def main() -> None:
    # Load events
    with EVENTS_CSV.open("r", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    for r in rows:
        r["price"] = float(r["price"]) if r["price"] else None
    print(f"Loaded {len(rows)} events from {EVENTS_CSV.name}", file=sys.stderr)

    rows = apply_manual_overrides(rows)
    print(f"After overrides: {len(rows)} events", file=sys.stderr)

    ledger = build_ledger(rows)
    open_n = sum(1 for st in ledger.values() if st["status"] == "OPEN")
    closed_n = sum(1 for st in ledger.values() if st["status"] == "CLOSED")
    print(f"Built ledger: {open_n} open / {closed_n} closed", file=sys.stderr)

    # Get unique entry/exit dates needed for SPY benchmark
    all_dates = set()
    for st in ledger.values():
        if st["first_entry_date"]:
            all_dates.add(st["first_entry_date"])
        if st["last_exit_date"]:
            all_dates.add(st["last_exit_date"])

    # Pull live quotes + SPY benchmarks
    print(f"Fetching live quotes for {len(ledger)} tickers...", file=sys.stderr)
    quotes: dict[str, float] = {}
    spy_entry: dict[str, float] = {}
    spy_exit: dict[str, float] = {}
    with httpx.Client() as client:
        for ticker in ledger:
            q = fetch_current_quote(client, ticker)
            if q:
                quotes[ticker] = q
        print(f"  got {len(quotes)} / {len(ledger)} quotes", file=sys.stderr)

        spy_current = fetch_current_quote(client, "SPY") or 0
        print(f"  SPY current: ${spy_current:.2f}", file=sys.stderr)

        print(f"Fetching SPY benchmark closes for {len(all_dates)} unique dates...",
              file=sys.stderr)
        spy_cache: dict[str, float] = {}
        for d in all_dates:
            if d not in spy_cache:
                p = fetch_close_on_date(client, "SPY", d)
                if p:
                    spy_cache[d] = p
        spy_entry = spy_cache.copy()
        spy_exit = spy_cache.copy()

    build_spreadsheet(ledger, quotes, spy_entry, spy_exit, spy_current)


if __name__ == "__main__":
    main()
