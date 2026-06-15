"""Wifey-swing-trades options ledger → Excel spreadsheet.

Reads parsed events CSV, builds a per-contract ledger (each unique
ticker+strike+exp+right = one trade), generates Excel matching the
discord-shared screenshot template:

  One row per EXIT EVENT (scale-out, close, stop, expire). Cost basis
  is the avg of all opens/adds prior to that exit. Open positions get
  one row with no exit price.

Columns:
  Status | SYM | Exp Date | Strike | Cost Avg | Exit Price | P/L % |
  Entry Date | Exit Date | Days Held | Note
"""
from __future__ import annotations

import csv
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
EVENTS_CSV = ROOT / "discord" / "wifey_parsed_events.csv"
OVERRIDES_CSV = ROOT / "discord" / "wifey_manual_overrides.csv"
TODAY = date.today().isoformat()
OUT_XLSX = ROOT / "discord" / f"wifey_swing_trades_{TODAY}.xlsx"

# ── Ledger ─────────────────────────────────────────────────────────────

def contract_key(e: dict) -> tuple | None:
    """Return (ticker, strike, expiration, right) or None if incomplete."""
    if not (e.get("ticker") and e.get("strike") and e.get("expiration") and e.get("right")):
        return None
    return (e["ticker"], float(e["strike"]), e["expiration"], e["right"])


def build_ledger(events: list[dict]) -> list[dict]:
    """Walk events per contract, build list of rows for spreadsheet.

    Each row represents ONE EXIT EVENT (or an open position with no exit).
    """
    # Group events by contract
    by_contract: dict[tuple, list[dict]] = defaultdict(list)
    for e in events:
        if e.get("is_spread") == "True":
            continue   # skip spreads for now
        if e["action"] == "OPEN_UNRESOLVED_EXP":
            continue   # skip — these are wrong-channel slip-throughs
        key = contract_key(e)
        if key is None:
            continue
        by_contract[key].append(e)

    rows: list[dict] = []
    for key, evs in by_contract.items():
        ticker, strike, exp, right = key
        evs.sort(key=lambda x: x["timestamp"])

        entries: list[tuple[str, float]] = []   # (date, price)
        units: float = 0.0
        peak_units: float = 0.0
        first_entry_date: str | None = None

        for e in evs:
            d = e["date"]
            try:
                p = float(e["price"]) if e["price"] else None
            except (ValueError, TypeError):
                p = None
            action = e["action"]

            if action in ("OPEN", "ADD"):
                if p is None:
                    continue
                if units == 0:
                    units = 1.0
                else:
                    units += 1.0
                entries.append((d, p))
                peak_units = max(peak_units, units)
                if first_entry_date is None:
                    first_entry_date = d

            elif action in ("TRIM",):
                if units <= 0 or p is None:
                    continue
                sold = units * 0.5
                avg_cost = sum(price for _, price in entries) / len(entries) if entries else 0
                pnl_pct = (p - avg_cost) / avg_cost * 100 if avg_cost else 0
                rows.append({
                    "status": "Closed",
                    "ticker": ticker,
                    "exp_date": exp,
                    "strike": f"{int(strike) if strike == int(strike) else strike}{right}",
                    "cost_avg": avg_cost,
                    "exit_price": p,
                    "pnl_pct": pnl_pct,
                    "entry_date": first_entry_date,
                    "exit_date": d,
                    "days_held": (datetime.fromisoformat(d).date()
                                  - datetime.fromisoformat(first_entry_date).date()).days
                                  if first_entry_date else None,
                    "note": f"Scale out 1/2 ({sold/peak_units:.0%} of peak)" if peak_units else "Scale out",
                })
                units -= sold

            elif action in ("CLOSE", "STOP", "EXPIRE"):
                if units <= 0:
                    continue
                avg_cost = sum(price for _, price in entries) / len(entries) if entries else 0
                if p is None and action == "EXPIRE":
                    p = 0.0   # expired worthless
                pnl_pct = (p - avg_cost) / avg_cost * 100 if avg_cost and p is not None else None
                note = {
                    "CLOSE": "Closed",
                    "STOP": "Stop loss",
                    "EXPIRE": "Expired",
                }.get(action, action)
                rows.append({
                    "status": "Closed",
                    "ticker": ticker,
                    "exp_date": exp,
                    "strike": f"{int(strike) if strike == int(strike) else strike}{right}",
                    "cost_avg": avg_cost,
                    "exit_price": p,
                    "pnl_pct": pnl_pct,
                    "entry_date": first_entry_date,
                    "exit_date": d,
                    "days_held": (datetime.fromisoformat(d).date()
                                  - datetime.fromisoformat(first_entry_date).date()).days
                                  if first_entry_date else None,
                    "note": note,
                })
                units = 0
                entries = []

            # ROLL events: treat as CLOSE of current contract + (separate
            # OPEN tracked elsewhere if explicit). For now, just close.
            # GUARD: multi-ticker summary messages like "Bought $GLD Rolled
            # $GOOGL" can wrongly assign ROLL to the first ticker. If the
            # ROLL has no explicit price (i.e. came from a summary, not
            # an exit announcement), skip it — don't close the position.
            elif action == "ROLL":
                if p is None:
                    continue   # phantom ROLL from multi-ticker message
                if units > 0 and entries:
                    avg_cost = sum(price for _, price in entries) / len(entries)
                    rows.append({
                        "status": "Closed",
                        "ticker": ticker,
                        "exp_date": exp,
                        "strike": f"{int(strike) if strike == int(strike) else strike}{right}",
                        "cost_avg": avg_cost,
                        "exit_price": p,
                        "pnl_pct": ((p - avg_cost) / avg_cost * 100) if (p and avg_cost) else None,
                        "entry_date": first_entry_date,
                        "exit_date": d,
                        "days_held": (datetime.fromisoformat(d).date()
                                      - datetime.fromisoformat(first_entry_date).date()).days
                                      if first_entry_date else None,
                        "note": "Rolled",
                    })
                    units = 0
                    entries = []

        # If contract still has units → check expiration
        if units > 0 and entries:
            avg_cost = sum(price for _, price in entries) / len(entries)
            # Auto-expire if expiration date is in the past (no explicit
            # close message in channel — wifey trades that decay naturally).
            try:
                exp_date = datetime.fromisoformat(exp).date()
                expired = exp_date < date.today()
            except (ValueError, TypeError):
                expired = False
            if expired:
                # Treat as expired worthless. P/L = -100% unless we have
                # evidence it expired ITM (no signal in data here, default
                # to worthless = conservative).
                rows.append({
                    "status": "Closed",
                    "ticker": ticker,
                    "exp_date": exp,
                    "strike": f"{int(strike) if strike == int(strike) else strike}{right}",
                    "cost_avg": avg_cost,
                    "exit_price": 0.0,
                    "pnl_pct": -100.0,
                    "entry_date": first_entry_date,
                    "exit_date": exp,
                    "days_held": ((exp_date
                                  - datetime.fromisoformat(first_entry_date).date()).days
                                  if first_entry_date else None),
                    "note": "Expired (no explicit close; assumed worthless)",
                })
            else:
                # Genuinely still open
                rows.append({
                    "status": "Open",
                    "ticker": ticker,
                    "exp_date": exp,
                    "strike": f"{int(strike) if strike == int(strike) else strike}{right}",
                    "cost_avg": avg_cost,
                    "exit_price": None,
                    "pnl_pct": None,
                    "entry_date": first_entry_date,
                    "exit_date": None,
                    "days_held": (date.today()
                                  - datetime.fromisoformat(first_entry_date).date()).days
                                  if first_entry_date else None,
                    "note": f"Open. {len(entries)} adds." if len(entries) > 1 else "Open.",
                })

    # Sort: open positions first, then closed by exit date desc
    rows.sort(key=lambda r: (
        r["status"] != "Open",
        -(datetime.fromisoformat(r["exit_date"]).timestamp()
          if r["exit_date"] else 9999999999),
    ))
    return rows


# ── Manual overrides ───────────────────────────────────────────────────

_NUMERIC_FIELDS = {"cost_avg", "exit_price", "pnl_pct", "days_held"}


def _parse_fields(blob: str) -> dict:
    """Parse 'key=value|key=value' override blob into a dict.

    Numeric fields are auto-coerced to float/int. Empty values left as-is.
    """
    out: dict = {}
    if not blob:
        return out
    for chunk in blob.split("|"):
        if "=" not in chunk:
            continue
        k, v = chunk.split("=", 1)
        k = k.strip()
        v = v.strip()
        if k in _NUMERIC_FIELDS and v:
            try:
                out[k] = float(v) if "." in v else int(v)
            except ValueError:
                out[k] = v
        else:
            out[k] = v if v else None
    return out


def _row_matches(row: dict, ticker: str, strike: str, exp_date: str,
                 exit_date: str) -> bool:
    """Match override key to a ledger row. Empty exit_date matches Open rows."""
    if (row.get("ticker") or "") != ticker:
        return False
    if (row.get("strike") or "") != strike:
        return False
    if (row.get("exp_date") or "") != exp_date:
        return False
    row_exit = row.get("exit_date") or ""
    return row_exit == exit_date


def apply_wifey_overrides(rows: list[dict]) -> list[dict]:
    """Apply operator-curated patches to ledger rows.

    Reads `discord/wifey_manual_overrides.csv` and performs DELETE / UPDATE /
    INSERT ops to align the auto-parsed ledger with the canonical screenshot
    ground truth (`discord/wifey_canonical_screenshot_2026-05-17.csv`).

    Override schema: op,ticker,strike,exp_date,exit_date,fields,note
      DELETE     remove matched row
      UPDATE     patch fields on matched row
      INSERT     create new row from match keys + fields
    """
    if not OVERRIDES_CSV.exists():
        return rows

    with OVERRIDES_CSV.open("r", encoding="utf-8") as f:
        ops = list(csv.DictReader(f))

    applied = {"DELETE": 0, "UPDATE": 0, "INSERT": 0, "skipped": 0}
    for op in ops:
        kind = (op.get("op") or "").strip()
        ticker = (op.get("ticker") or "").strip()
        strike = (op.get("strike") or "").strip()
        exp_date = (op.get("exp_date") or "").strip()
        exit_date = (op.get("exit_date") or "").strip()
        fields = _parse_fields(op.get("fields") or "")

        if kind == "DELETE":
            before = len(rows)
            rows = [r for r in rows
                    if not _row_matches(r, ticker, strike, exp_date, exit_date)]
            if len(rows) < before:
                applied["DELETE"] += 1
            else:
                applied["skipped"] += 1
                print(f"  [override-skip] DELETE no match: {ticker} {strike} "
                      f"{exp_date} exit={exit_date!r}", file=sys.stderr)

        elif kind == "UPDATE":
            matched = False
            for r in rows:
                if _row_matches(r, ticker, strike, exp_date, exit_date):
                    for k, v in fields.items():
                        r[k] = v
                    matched = True
            if matched:
                applied["UPDATE"] += 1
            else:
                applied["skipped"] += 1
                print(f"  [override-skip] UPDATE no match: {ticker} {strike} "
                      f"{exp_date} exit={exit_date!r}", file=sys.stderr)

        elif kind == "INSERT":
            new_row: dict = {
                "status": fields.get("status", "Open"),
                "ticker": ticker,
                "exp_date": exp_date,
                "strike": strike,
                "cost_avg": fields.get("cost_avg"),
                "exit_price": fields.get("exit_price"),
                "pnl_pct": fields.get("pnl_pct"),
                "entry_date": fields.get("entry_date"),
                "exit_date": exit_date or None,
                "days_held": fields.get("days_held"),
                "note": fields.get("note", ""),
            }
            rows.append(new_row)
            applied["INSERT"] += 1

        else:
            applied["skipped"] += 1

    print(f"Applied overrides: {applied}", file=sys.stderr)

    # Re-sort after patches: open first, then closed by exit date desc.
    rows.sort(key=lambda r: (
        r["status"] != "Open",
        -(datetime.fromisoformat(r["exit_date"]).timestamp()
          if r["exit_date"] else 9999999999),
    ))
    return rows


# ── Spreadsheet ────────────────────────────────────────────────────────

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
BODY_FONT = Font(name="Arial")
OPEN_FILL = PatternFill("solid", start_color="C6EFCE")  # green
CLOSED_FILL = PatternFill("solid", start_color="FFC7CE")  # red
GAIN_FILL_BRIGHT = PatternFill("solid", start_color="00B050")
GAIN_FILL_LIGHT = PatternFill("solid", start_color="C6EFCE")
LOSS_FILL = PatternFill("solid", start_color="FFC7CE")
LOSS_FILL_DEEP = PatternFill("solid", start_color="FF0000")


def pnl_fill(pnl: float | None) -> PatternFill | None:
    if pnl is None:
        return None
    if pnl >= 50:
        return GAIN_FILL_BRIGHT
    if pnl > 0:
        return GAIN_FILL_LIGHT
    if pnl <= -50:
        return LOSS_FILL_DEEP
    return LOSS_FILL


def build_spreadsheet(rows: list[dict], perf: list[dict] | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wifey Trades"

    headers = ["Status", "SYM", "Exp Date", "Strike", "Cost Avg",
               "Exit Price", "P/L %", "Entry Date", "Exit Date",
               "Days Held", "Note"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for r in rows:
        row_data = [
            r["status"], r["ticker"], r["exp_date"], r["strike"],
            r["cost_avg"], r["exit_price"],
            (r["pnl_pct"] / 100) if r["pnl_pct"] is not None else None,
            r["entry_date"], r["exit_date"], r["days_held"], r["note"],
        ]
        ws.append(row_data)
        row_idx = ws.max_row

        # Status fill
        ws.cell(row=row_idx, column=1).fill = (
            OPEN_FILL if r["status"] == "Open" else CLOSED_FILL
        )

        # P/L fill
        fill = pnl_fill(r["pnl_pct"])
        if fill:
            ws.cell(row=row_idx, column=7).fill = fill

        for c in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=c).font = BODY_FONT
            ws.cell(row=row_idx, column=c).border = BORDER

    # Number formats
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=5).number_format = '$#,##0.00'  # Cost
        ws.cell(row=row_idx, column=6).number_format = '$#,##0.00'  # Exit
        ws.cell(row=row_idx, column=7).number_format = '0.00%'       # P/L
        ws.cell(row=row_idx, column=10).number_format = '0'           # Days

    # Column widths
    widths = [8, 7, 12, 9, 10, 11, 10, 12, 12, 7, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    if perf:
        build_perf_tab(wb, perf)

    target = OUT_XLSX
    try:
        wb.save(target)
    except PermissionError:
        # File open in Excel — rotate to versioned name so we don't lose the run.
        for v in range(2, 99):
            alt = target.with_name(f"{target.stem}_v{v}.xlsx")
            if not alt.exists():
                target = alt
                break
        wb.save(target)
        print(f"  [warn] base file locked; saved as {target.name}", file=sys.stderr)
    print(f"Wrote spreadsheet to {target}", file=sys.stderr)


# ── Performance tab (cumulative return vs SPY) ─────────────────────────

PERF_ANCHOR_DATE = "2025-01-02"  # SPY history start; both series start at 0%


def _fetch_spy_daily() -> list[tuple[date, float]]:
    """Pull SPY daily closes from PERF_ANCHOR_DATE through today via Tradier.

    Returns list of (date, close) sorted ascending. Returns [] on failure
    so the rest of the spreadsheet still builds.
    """
    try:
        import asyncio
        if str(ROOT) not in sys.path:
            sys.path.insert(0, str(ROOT))
        from server.tradier import TradierClient

        async def _go():
            c = TradierClient()
            try:
                bars = await c.history(
                    "SPY", interval="daily",
                    start=PERF_ANCHOR_DATE,
                    end=date.today().isoformat(),
                )
            finally:
                await c.close()
            return bars

        bars = asyncio.run(_go())
        out: list[tuple[date, float]] = []
        for b in bars:
            try:
                d = datetime.fromisoformat(b["time"]).date()
                out.append((d, float(b["close"])))
            except (KeyError, ValueError, TypeError):
                continue
        out.sort(key=lambda x: x[0])
        return out
    except Exception as e:
        print(f"  [perf] SPY fetch failed: {e}", file=sys.stderr)
        return []


def _build_perf_series(rows: list[dict], spy_bars: list[tuple[date, float]]) -> list[dict]:
    """Build a per-calendar-day performance series.

    For each SPY trading day:
      - wifey_pnl_dollars  = cumulative $ P&L assuming ONE CONTRACT per trade.
                             $ P&L per trade = (exit_price - cost_avg) * 100.
                             Sizing-agnostic ground truth (the wifey's actual
                             sizing is unknown — this is the unweighted ledger).
                             Avoids absorbing-state of multiplicative
                             compounding when any single trade is -100%
                             (expired-worthless options).
      - spy_cum            = (close / anchor_close) - 1   [SPY buy-and-hold]

    Two metrics, two axes. Reader can interpret edge envelope vs market.
    """
    if not spy_bars:
        return []
    anchor_close = spy_bars[0][1]

    # Build sorted list of (exit_date, $ P&L per contract) for closed rows.
    exits: list[tuple[date, float]] = []
    for r in rows:
        if r.get("status") != "Closed":
            continue
        ex = r.get("exit_date")
        cost = r.get("cost_avg")
        exit_p = r.get("exit_price")
        if ex is None or cost is None or exit_p is None:
            continue
        try:
            d = datetime.fromisoformat(ex).date()
        except (ValueError, TypeError):
            continue
        if d < spy_bars[0][0]:
            continue
        pnl_dollars = (float(exit_p) - float(cost)) * 100.0
        exits.append((d, pnl_dollars))
    exits.sort(key=lambda x: x[0])

    out: list[dict] = []
    cum_dollars = 0.0
    j = 0
    n_trades = 0
    for d, close in spy_bars:
        while j < len(exits) and exits[j][0] <= d:
            cum_dollars += exits[j][1]
            n_trades += 1
            j += 1
        out.append({
            "date": d,
            "wifey_pnl_dollars": cum_dollars,
            "spy_cum": (close / anchor_close) - 1.0,
            "trades_to_date": n_trades,
        })
    # Stash per-trade $ P&L list on the first row for downstream stats.
    if out:
        out[0]["_per_trade_pnls"] = [p for _, p in exits]
    return out


def build_perf_tab(wb: Workbook, perf: list[dict]) -> None:
    """Add a Performance sheet with the cumulative-return table + line chart."""
    if not perf:
        print("  [perf] no data — skipping Performance tab", file=sys.stderr)
        return
    ws = wb.create_sheet("Performance")

    # Headline summary
    final = perf[-1]
    # Per-trade $ P&L stats (true trade-level, not day-aggregated)
    pnls: list[float] = perf[0].get("_per_trade_pnls", [])
    wins = sum(1 for p in pnls if p > 0)
    losses = sum(1 for p in pnls if p < 0)
    breakeven = sum(1 for p in pnls if p == 0)
    n_closed = wins + losses
    win_rate = (wins / n_closed * 100) if n_closed else 0.0
    avg_win = (sum(p for p in pnls if p > 0) / wins) if wins else 0.0
    avg_loss = (sum(p for p in pnls if p < 0) / losses) if losses else 0.0
    total_wins = sum(p for p in pnls if p > 0)
    total_losses = abs(sum(p for p in pnls if p < 0))
    profit_factor = (total_wins / total_losses) if total_losses > 0 else float("inf")

    ws["A1"] = "Wifey Swing Trades — Equal-Weight Performance vs SPY"
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws["A2"] = f"Window: {perf[0]['date'].isoformat()} → {final['date'].isoformat()}"
    ws["A3"] = (f"1-contract cumulative P&L: ${final['wifey_pnl_dollars']:+,.0f}   "
                f"|   SPY buy-and-hold: {final['spy_cum']*100:+.1f}%   "
                f"|   {final['trades_to_date']} closed trades")
    be_note = f" / {breakeven}BE" if breakeven else ""
    ws["A4"] = (f"Win rate: {win_rate:.1f}% ({wins}W / {losses}L{be_note})   "
                f"|   Avg win: ${avg_win:+,.0f}   "
                f"|   Avg loss: ${avg_loss:+,.0f}   "
                f"|   Profit factor: {profit_factor:.2f}")
    for row, sz in ((1, 14), (2, 11), (3, 11), (4, 11)):
        ws.cell(row=row, column=1).font = Font(
            bold=(row == 1), name="Arial", size=sz
        )

    # Table headers at row 6
    headers = ["Date", "Wifey Cum $ P&L (1ct)", "SPY Cum Return", "Trades To-Date"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    start_row = 7
    for i, p in enumerate(perf):
        r = start_row + i
        ws.cell(row=r, column=1, value=p["date"]).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=p["wifey_pnl_dollars"]).number_format = '$#,##0;($#,##0);-'
        ws.cell(row=r, column=3, value=p["spy_cum"]).number_format = "0.0%"
        ws.cell(row=r, column=4, value=p["trades_to_date"]).number_format = "0"
        for c in range(1, 5):
            ws.cell(row=r, column=c).font = BODY_FONT

    end_row = start_row + len(perf) - 1

    # Column widths
    for col, w in zip("ABCD", [14, 22, 18, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A7"

    # Chart — dual axis: wifey $ on primary, SPY % on secondary
    chart = LineChart()
    chart.title = "Equal-Weight 1-Contract P&L vs SPY Buy-and-Hold"
    chart.y_axis.title = "Wifey Cum $ P&L (per contract)"
    chart.x_axis.title = "Date"
    chart.y_axis.number_format = '"$"#,##0'
    chart.height = 12
    chart.width = 24
    chart.legend.position = "b"

    # Series 1: wifey $
    data1 = Reference(ws, min_col=2, max_col=2,
                      min_row=6, max_row=end_row)
    chart.add_data(data1, titles_from_data=True)

    # Series 2: SPY % — needs a second axis
    spy_chart = LineChart()
    data2 = Reference(ws, min_col=3, max_col=3,
                     min_row=6, max_row=end_row)
    spy_chart.add_data(data2, titles_from_data=True)
    spy_chart.y_axis.axId = 200
    spy_chart.y_axis.crosses = "max"
    spy_chart.y_axis.title = "SPY Cum Return"
    spy_chart.y_axis.number_format = "0.0%"

    cats = Reference(ws, min_col=1, max_col=1,
                     min_row=7, max_row=end_row)
    chart.set_categories(cats)

    # Merge SPY series into main chart on secondary axis
    chart += spy_chart

    ws.add_chart(chart, "F6")


# ── Main ────────────────────────────────────────────────────────────────

def main() -> None:
    with EVENTS_CSV.open("r", encoding="utf-8") as f:
        events = list(csv.DictReader(f))
    print(f"Loaded {len(events)} events", file=sys.stderr)

    rows = build_ledger(events)
    print(f"Built {len(rows)} ledger rows", file=sys.stderr)

    rows = apply_wifey_overrides(rows)
    print(f"After overrides: {len(rows)} ledger rows", file=sys.stderr)

    print("Fetching SPY daily history...", file=sys.stderr)
    spy_bars = _fetch_spy_daily()
    perf = _build_perf_series(rows, spy_bars)
    print(f"Built perf series: {len(perf)} days, {len(spy_bars)} SPY bars",
          file=sys.stderr)

    open_n = sum(1 for r in rows if r["status"] == "Open")
    closed_n = sum(1 for r in rows if r["status"] == "Closed")
    print(f"  Open positions: {open_n}", file=sys.stderr)
    print(f"  Closed exit rows: {closed_n}", file=sys.stderr)

    build_spreadsheet(rows, perf=perf)


if __name__ == "__main__":
    main()
